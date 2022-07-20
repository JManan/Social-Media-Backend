from decouple import config
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import UserRegisterForm
from django.contrib.auth.decorators import login_required
from .forms import UserUpdateForm, ProfileUpdateForm
from .models import Profile
from django.contrib.auth.mixins import  LoginRequiredMixin
from django.views import generic
from django.views import View
from app.models import PostList
from openpyxl import Workbook
from django.http import HttpResponse


def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'Account created for {username}!')
            return redirect('index')
    else:
        form = UserRegisterForm()
    return render(request, 'user/register.html', {'form' : form})

@login_required
def profile(request):
    if request.method == 'POST':
        u_form = UserUpdateForm(request.POST , instance=request.user)
        p_form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user.profile)
        if u_form.is_valid() and p_form.is_valid():
            u_form.save()
            p_form.save()
            username = u_form.cleaned_data.get('username')
            messages.success(request, f'Changes Saved for {username}! profile')
            return redirect('profile')
    else:   
        u_form = UserUpdateForm(instance=request.user)
        p_form = ProfileUpdateForm(instance=request.user.profile)
        context = {
        'u_form' : u_form,
        'p_form' : p_form,
    }
        return render(request, 'user/profile.html', context)


class ProfileDetailView(LoginRequiredMixin, generic.DetailView):
    model = Profile
    # context_object_name = 'profile'
    # template_name = 'user/other-profile.html'

    

    def get(self, request, pk, *args, **kwargs):
        profile = Profile.objects.get(user_id=pk)
        user = profile.user
        # posts = PostList.objects.filter(author=user).order_by('-date_pb')

        followers = profile.followers.all()

        if len(followers) == 0:
            is_following = False


        for follower in followers:
            if follower == request.user:
                is_following = True
                break
            else:
                is_following = False

        number_of_followers = len(followers)

        context = {
            'user' : user,
            'profile' : profile,
            'number_of_followers' : number_of_followers,
            'is_following' : is_following
        }
    
        return render(request, 'user/other-profile.html', context)

class AddFollower(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        profile = Profile.objects.get(pk=pk)
        print(profile.followers.all, request.user)
        profile.followers.add(request.user)
        print(profile.followers.all, request.user)
        return redirect('other-profile', profile.pk)

class RemoveFollower(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        profile = Profile.objects.get(pk=pk)
        profile.followers.remove(request.user)
        return redirect('other-profile', profile.pk)

def profiles_in_xlxs(request):
    profile_queryset = Profile.objects.all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Profiles'
    row_num = 0
    columns = [
        'ID',
        'User',
        'Age',
        'Address',
        'Number',
    ]

    for col_num in range(len(columns)):
        c = worksheet.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num]
    
    row_num = 1

    for obj in profile_queryset:
        row = [
        str(obj.pk),
        str(obj.user),
        str(obj.address),
        str(obj.age),
        str(obj.number),
        ]

        for col_num in range(len(row)):
            c = worksheet.cell(row=row_num+1, column=col_num+1)
            c.value = row[col_num]
        row_num+=1

    workbook.save(response)
    return response


class Add_Email_Receiver(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        profile = Profile.objects.get(pk=pk)
        # print(profile.followers.all, request.user)
        profile.followers.add(request.user)
        a = profile.email_receivers.add(self.request.user)
        print(a)
        profile.email_receivers.add(self.request.user)
        print(profile.email_receivers.all(),profile.followers.all(), request.user)
        print(type(profile.email_receivers))
        # print(a)
        # print(2)
        return redirect('other-profile', profile.pk)

class Remove_Email_Receiver(LoginRequiredMixin, View):
    model = Profile
    def post(self, request, pk, *args, **kwargs):
        profile = Profile.objects.get(pk=pk)
        profile.followers.remove(request.user)
        profile.email_receivers.remove(request.user)
        return redirect('other-profile', profile.pk)
